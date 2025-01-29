from fastapi import APIRouter, Request, Form, HTTPException
import boto3
import pandas as pd
from io import StringIO
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import requests
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


# router = APIRouter()

# # S3 Configuration
# BUCKET_NAME = "glacier-ml-training"
# METADATA_KEY = "universal-db/metadata.csv"
# OUTPUT_FOLDER = "review-results/"
# RESULTS_FILE = f"{OUTPUT_FOLDER}results.csv"
# s3_client = boto3.client("s3")

# # Temporary storage for session reviews
# session_reviews = []


# def save_results_excel_with_images():
#     """
#     Save the results as an Excel file with embedded images, ensuring consistent checks as with the CSV creation process.
#     """
#     try:
#         # Define the local path for the Excel file
#         excel_path = "/tmp/review_results.xlsx"

#         # Fetch the existing results from S3
#         print(f"Fetching {RESULTS_FILE} from S3...")
#         results_df = get_results_df()

#         # Drop duplicates based on "id" to ensure unique rows
#         results_df = results_df.drop_duplicates(subset="id")
#         print(f"Filtered unique results: {len(results_df)} rows.")

#         # Create a new Excel workbook
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Review Results"
#         headers = [
#             "id",
#             "color",
#             "brand",
#             "material",
#             "shape",
#             "modifiers",
#             "reviewed_by",
#             "errors",
#             "image",
#         ]
#         ws.append(headers)

#         # Set column width and row height
#         ws.column_dimensions[get_column_letter(9)].width = 25  # "image" column width
#         row_height = 100  # Set row height for image visibility

#         # Add rows and embed images
#         for _, row in results_df.iterrows():
#             # Generate presigned URL for the current row
#             image_url = generate_presigned_url(row["s3_file_path"])
#             image_path = None

#             if image_url:
#                 try:
#                     # Fetch the image from the presigned URL
#                     response = requests.get(image_url)
#                     if response.status_code == 200:
#                         image = PILImage.open(BytesIO(response.content))
#                         image_path = f"/tmp/{row['id']}.png"
#                         image.save(image_path)
#                         print(f"Image saved temporarily: {image_path}")
#                     else:
#                         print(
#                             f"Failed to download image: {image_url} (Status Code: {response.status_code})"
#                         )
#                 except Exception as e:
#                     print(f"Error downloading image: {str(e)}")

#             # Append data row to the Excel sheet
#             ws.append(
#                 [
#                     row.get("id"),
#                     row.get("color"),
#                     row.get("brand"),
#                     row.get("material"),
#                     row.get("shape"),
#                     row.get("modifiers"),
#                     row.get("reviewed_by"),
#                     (
#                         ", ".join(eval(row["errors"]))
#                         if isinstance(row["errors"], str)
#                         else ""
#                     ),
#                     None,  # Placeholder for the image
#                 ]
#             )

#             # Embed the image in the Excel file
#             if image_path and os.path.exists(image_path):
#                 try:
#                     excel_image = ExcelImage(image_path)
#                     excel_image.height = 80  # Adjust image height
#                     excel_image.width = 80  # Adjust image width
#                     ws.add_image(
#                         excel_image, f"I{ws.max_row}"
#                     )  # Embed image in column "I"
#                     print(f"Image embedded in Excel at row {ws.max_row}.")
#                 except Exception as e:
#                     print(f"Error embedding image in Excel: {str(e)}")
#             else:
#                 print(f"Image path does not exist or is invalid: {image_path}")

#             # Adjust the row height for the current row
#             ws.row_dimensions[ws.max_row].height = row_height

#         # Save the Excel file locally
#         wb.save(excel_path)
#         print(f"Excel file saved locally: {excel_path}")

#         # Upload the updated Excel file to S3
#         with open(excel_path, "rb") as f:
#             s3_client.put_object(
#                 Bucket=BUCKET_NAME, Key=f"{OUTPUT_FOLDER}review_results.xlsx", Body=f
#             )
#         print("Excel file with images uploaded to S3.")

#     except Exception as e:
#         print(f"Error saving Excel with images: {str(e)}")
#         raise HTTPException(status_code=500, detail="Error saving Excel with images")


# def get_results_df():
#     """
#     Fetch the existing results CSV from S3. If it doesn't exist, return an empty DataFrame.
#     """
#     try:
#         print(f"Checking for {RESULTS_FILE} in S3...")
#         obj = s3_client.get_object(Bucket=BUCKET_NAME, Key=RESULTS_FILE)
#         results_csv = obj["Body"].read().decode("utf-8")
#         print(f"{RESULTS_FILE} fetched successfully.")
#         return pd.read_csv(StringIO(results_csv))
#     except s3_client.exceptions.NoSuchKey:
#         print(f"{RESULTS_FILE} does not exist. Creating a new DataFrame.")
#         return pd.DataFrame(
#             columns=[
#                 "id",
#                 "color",
#                 "brand",
#                 "material",
#                 "shape",
#                 "modifiers",
#                 "reviewed_by",
#                 "errors",
#             ]
#         )
#     except Exception as e:
#         print(f"Error fetching {RESULTS_FILE}: {str(e)}")
#         raise HTTPException(status_code=500, detail="Error fetching results CSV")


# def save_results_df(new_entries_df):
#     """
#     Fetch the existing results DataFrame, append new entries, and save back to S3.
#     """
#     try:
#         # Fetch the existing results
#         existing_df = get_results_df()
#         print(f"Existing results loaded: {len(existing_df)} rows.")

#         # Append new entries
#         updated_df = pd.concat([existing_df, new_entries_df], ignore_index=True)
#         print(f"Updated results: {len(updated_df)} rows.")

#         # Save the updated DataFrame back to S3
#         csv_buffer = StringIO()
#         updated_df.to_csv(csv_buffer, index=False)
#         s3_client.put_object(
#             Bucket=BUCKET_NAME, Key=RESULTS_FILE, Body=csv_buffer.getvalue()
#         )
#         print(f"{RESULTS_FILE} updated successfully.")
#     except Exception as e:
#         print(f"Error saving results CSV: {str(e)}")
#         raise HTTPException(status_code=500, detail="Error saving results CSV")


# def append_to_csv_s3(field, new_data):
#     """
#     Append new data to an existing CSV in S3 or create a new one if it doesn't exist.
#     """
#     try:
#         # Define the S3 path for the CSV
#         field_csv_path = f"{OUTPUT_FOLDER}{field}_errors.csv"

#         # Fetch existing data if the file exists
#         try:
#             obj = s3_client.get_object(Bucket=BUCKET_NAME, Key=field_csv_path)
#             existing_csv = obj["Body"].read().decode("utf-8")
#             existing_df = pd.read_csv(StringIO(existing_csv))
#             print(f"{field_csv_path} fetched successfully. {len(existing_df)} rows.")
#         except s3_client.exceptions.NoSuchKey:
#             print(f"{field_csv_path} does not exist. Creating a new DataFrame.")
#             existing_df = pd.DataFrame()

#         # Append new data
#         updated_df = pd.concat([existing_df, new_data], ignore_index=True)

#         # Save the updated DataFrame back to S3
#         csv_buffer = StringIO()
#         updated_df.to_csv(csv_buffer, index=False)
#         s3_client.put_object(
#             Bucket=BUCKET_NAME, Key=field_csv_path, Body=csv_buffer.getvalue()
#         )
#         print(f"{field_csv_path} updated successfully. {len(updated_df)} rows.")
#     except Exception as e:
#         print(f"Error appending to {field_csv_path}: {str(e)}")
#         raise HTTPException(status_code=500, detail=f"Error updating {field_csv_path}")


# def generate_presigned_url(s3_path):
#     """
#     Generate a presigned URL for an S3 file.
#     """
#     path_without_prefix = s3_path.replace("s3://", "")
#     bucket, key = path_without_prefix.split("/", 1)
#     url = s3_client.generate_presigned_url(
#         "get_object", Params={"Bucket": bucket, "Key": key}, ExpiresIn=3600
#     )
#     return url


# @router.get("/review/metadata")
# async def review_metadata(image_index: int = 0):
#     """
#     Fetch metadata and generate presigned URL for a given image index.
#     Only process unreviewed rows with status == "active".
#     """
#     print(f"Request received for image_index: {image_index}")

#     try:
#         # Fetch metadata.csv from S3
#         print("Fetching metadata.csv from S3...")
#         obj = s3_client.get_object(Bucket=BUCKET_NAME, Key=METADATA_KEY)
#         metadata_csv = obj["Body"].read().decode("utf-8")
#         print("CSV content decoded successfully.")
#         df = pd.read_csv(StringIO(metadata_csv))
#         df = df[df["status"] == "active"]
#         print(f"Filtered active rows. Number of rows: {len(df)}.")

#         # Exclude reviewed images
#         results_df = get_results_df()
#         reviewed_ids = set(results_df["id"].tolist())
#         df = df[~df["id"].isin(reviewed_ids)]
#         print(f"Filtered unreviewed rows. Number of rows: {len(df)}.")

#         # Check index validity
#         if image_index >= len(df):
#             print("Image index out of range. Returning completion status.")
#             return {"status": "complete", "message": "All images reviewed"}

#         # Fetch the row for the given index
#         row = df.iloc[image_index]
#         print(f"Row fetched: {row.to_dict()}")

#         # Generate presigned URL for the image
#         image_url = generate_presigned_url(row["s3_file_path"])
#         print(f"Presigned URL generated: {image_url}")

#         # Sanitize metadata to remove NaN or invalid float values
#         metadata = row.to_dict()
#         sanitized_metadata = {
#             key: (
#                 value
#                 if key in ["brand", "color", "shape", "modifier", "material"]
#                 and value is not None
#                 and pd.notna(value)
#                 and not (isinstance(value, (int, float)) and np.isinf(value))
#                 else None
#             )
#             for key, value in metadata.items()
#         }
#         print(f"Sanitized metadata: {sanitized_metadata}")

#         # Return success response
#         return {
#             "status": "success",
#             "image_index": image_index,
#             "total_images": len(df),
#             "image_url": image_url,
#             "metadata": sanitized_metadata,
#         }

#     except Exception as e:
#         print(f"Error in review_metadata: {str(e)}")
#         raise HTTPException(
#             status_code=500, detail=f"Error fetching metadata: {str(e)}"
#         )


# @router.post("/review/metadata")
# async def submit_review(
#     image_index: int = Form(...),
#     incorrect_color: bool = Form(False),
#     incorrect_brand: bool = Form(False),
#     incorrect_material: bool = Form(False),
#     incorrect_shape: bool = Form(False),
#     incorrect_modifiers: bool = Form(False),
# ):
#     """
#     Submit review and temporarily store results in session_reviews.
#     """
#     try:
#         print(f"Submitting review for image_index: {image_index}")

#         # Check if the image has already been reviewed in the current session
#         if any(review["id"] == image_index for review in session_reviews):
#             print(f"Image {image_index} already reviewed in this session. Skipping.")
#             # Return the next image index without processing the current one
#             return {"next_image": image_index + 1}

#         # Fetch metadata.csv
#         obj = s3_client.get_object(Bucket=BUCKET_NAME, Key=METADATA_KEY)
#         metadata_csv = obj["Body"].read().decode("utf-8")
#         df = pd.read_csv(StringIO(metadata_csv))
#         df = df[df["status"] == "active"]

#         # Exclude reviewed images
#         results_df = get_results_df()
#         reviewed_ids = set(results_df["id"].tolist())
#         df = df[~df["id"].isin(reviewed_ids)]

#         # Validate image index
#         if image_index >= len(df):
#             raise HTTPException(status_code=400, detail="Invalid image index")

#         # Get the current row
#         row = df.iloc[image_index]

#         # Track incorrect fields
#         incorrect_fields = []
#         if incorrect_color:
#             incorrect_fields.append("color")
#         if incorrect_brand:
#             incorrect_fields.append("brand")
#         if incorrect_material:
#             incorrect_fields.append("material")
#         if incorrect_shape:
#             incorrect_fields.append("shape")
#         if incorrect_modifiers:
#             incorrect_fields.append("modifiers")

#         # Append review to session_reviews
#         print(f"Session reviews at adding: {session_reviews}")

#         session_reviews.append(
#             {
#                 "id": row["id"],
#                 "color": row["color"],
#                 "brand": row["brand"],
#                 "material": row["material"],
#                 "shape": row["shape"],
#                 "modifiers": row["modifier"],
#                 "reviewed_by": os.getenv("REVIEWER_NAME", "anonymous"),
#                 "errors": incorrect_fields,
#                 "s3_file_path": row["s3_file_path"],
#             }
#         )
#         return {"next_image": image_index + 1}

#     except Exception as e:
#         print(f"Error in submit_review: {str(e)}")
#         raise HTTPException(
#             status_code=500, detail=f"Error submitting review: {str(e)}"
#         )


# @router.post("/review/metadata/complete")
# async def complete_review():
#     """
#     Save results from session_reviews into separate CSV files, create Excel with images, and upload to S3.
#     """
#     try:
#         global session_reviews
#         if not session_reviews:
#             return {"status": "success", "message": "No reviews to complete."}

#         # Convert session_reviews to DataFrame
#         reviews_df = pd.DataFrame(session_reviews)

#         # Save to separate CSVs for incorrect fields
#         for field in ["color", "brand", "material", "shape", "modifiers"]:
#             incorrect_df = reviews_df[reviews_df["errors"].apply(lambda x: field in x)]
#             if not incorrect_df.empty:
#                 append_to_csv_s3(field, incorrect_df)

#         # Save the main results CSV
#         save_results_df(reviews_df)

#         # Save results as an Excel file with images
#         save_results_excel_with_images()

#         # Clear session data
#         session_reviews = []

#         return {"status": "success", "message": "Review process completed."}

#     except Exception as e:
#         print(f"Error in complete_review: {str(e)}")
#         raise HTTPException(
#             status_code=500, detail=f"Error completing review: {str(e)}"
#         )
